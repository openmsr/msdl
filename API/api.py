import os
import asdf
import numpy as np
import matplotlib.pyplot as plt
import math
import re
from scipy import optimize
import openpyxl

#For Kriging model
import pyKriging  
from pyKriging.krige import kriging  
from pyKriging.samplingplan import samplingplan
from openpyxl.drawing.image import Image
from pyKriging.regressionkrige import regression_kriging

#For the surface plot
from matplotlib import cm
from matplotlib.ticker import LinearLocator, FormatStrFormatter
from mpl_toolkits.mplot3d import Axes3D

class API:
    def __init__(self, propert, salts, composition = []):
        self.propert = propert         #Desired physical property
        self.salts = salts             #List of desired salt combo
        self.composition = composition #Salt composition
        self.regressionType = 'None' #Description of the Regression type
        self.modelType = 'Regression kriging'    #Type of model to fit the data (Default Kriging for now)
        self.fitModel = 'None'      #Objec to store the fit model
        self.source = []            #List of asdf files that include the salt combo
        self.allSets = []          #List of all datasets currently in library
        
        self.loc = os.getcwd()
        
        #Setting the unit of the physical property (For picture caption)
        if propert.lower() == 'viscosity':
            self.unit = 'Pa*s'
        elif propert.lower() == 'density':
            self.unit = 'g/cm3'
        elif propert.lower() == 'thermal conductivity':
            self.unit = 'W/(m K)'
        elif propert.lower() == 'surface tension':
            self.unit = 'dyn/cm'
        elif propert.lower() == 'electric conductance':
            self.unit = '1/(ohm cm)'
        elif propert.lower() == 'heat capacity':
            self.unit = 'cal/(K mol)'
        elif propert.lower() == 'vapor pressure':
            self.unit = 'mmHg'
        else:
            self.unit = 'Unit undefined'
            
    #Scan the database library and extract datasets that contain the desired salts
    def scanLibrary(self):
        dataSetList = []
        #Get a list of all asdf files in Library and loop through them
        apiLoc = os.getcwd()
        
        os.chdir('..\\Library')
        
        for file in os.listdir():
            if not file.endswith('asdf'):   #Only look through .asdf files
                continue
            af = asdf.open(file)
            
            for dataSet in af.tree:
                if not dataSet.startswith('dataSet'):   #Skip searching the metadata tree
                    continue
                
                
                #Ensuring that we can directly compare salt list by lowercasing and sorting them
                saltList = [a.lower() for a in af.tree[dataSet]['Components']] 
                self.allSets.append(saltList)
                requestList = [a.lower() for a in self.salts]
                saltList.sort()
                requestList.sort()
                
                #Only work with files that match the specified salt combo
                if saltList == requestList:
                    if self.composition == []:
                        dataSetList.append(af.tree[dataSet])
                        self.source.append(file)    #Append filename to the source list
                    else:
                        
                        if all(np.array(self.composition) == np.array(af.tree[dataSet]['Composition'])):     #If composition is specified, only grab specific composition
                            dataSetList.append(af.tree[dataSet])
                            self.source.append(file)
                        else:
                            continue
                
                
        
        #Finally change the working directory back to the original one
        os.chdir(apiLoc)
        self.dataSetList = dataSetList  #Storing raw data in API object
        self.source = set(self.source)  #Delete source duplicates
        
        return dataSetList
    
    #Organizes all the salt measurements into X and Y, where X is a matrix where each column represents composition and temperature
    #and Y contains the property measurements
    def organizeData(self):
        numSalts = len(self.salts)
        l = 0
        for dataSet in self.dataSetList:
            if dataSet['Property'][0].lower() == self.propert.lower():
                
                #Regular expression for asdf keys (Attempt to make more robust in case misspelling in excel files occurr)
                allKeys = list(dataSet.keys())
                allKeysString = ' '.join(allKeys)
                tempRegex = re.compile(r'temp[\S]*|Temp[\S]*')
                mo1 = tempRegex.search(allKeysString)
                tempKey = mo1.group()
                measureRegex = re.compile(r'meas[\S]*|Meas[\S]*')
                mo1 = measureRegex.search(allKeysString)
                measureKey = mo1.group()
                compRegex = re.compile(r'comp[\S]*ion|Comp[\S]*ion')
                mo1 = compRegex.search(allKeysString)
                compKey = mo1.group()
                
                
                temps = np.array(dataSet[tempKey])  #Sometimes temperature arrays don't come out as numpy arrays (If they only have integers), this ensures it
                
                #Creating X and Y vectors
                tempMat = np.zeros( (numSalts + 1 ,temps.shape[0]))
                for i in range(numSalts):
                    tempMat[i,:] = np.repeat(dataSet[compKey][i],temps.shape[0])    #One composition is measured many times, so we create a proper array size to match temperature readings
                tempMat[-1,:] = temps
                if l == 0:
                    X = tempMat
                    Y = dataSet[measureKey]
                else:
                    X = np.concatenate((X,tempMat),axis=1)
                    Y = np.concatenate((Y,dataSet[measureKey]))
                l = l+1
        
        #Do to keeping track with variable l, if it's 0 it means no datasets were found that match the physical property        
        if l>0:        
            self.X = X
            self.Y = np.array(Y)
        #Set flags and ending function early to prevent raising exceptions
        else:
            print('No data in libray matches specifications')
            self.oneDimensional = False
            self.binary = False
            return 0    
        
        #Checking if salt data contains multiple composition, in the case it will do a one dimensional regression
        self.oneDimensional = True
        for i in range(X.shape[0]-1):
            oneComp = X[i,:]
            checkers = (oneComp[1:] == oneComp[:-1])
            if all(checkers):
                pass
            else:
                self.oneDimensional = False
                
                
        #Check for binary mixture (only relevant for multiple compositions)
        if not self.oneDimensional and X.shape[0] == 3:
            self.binary = True
            #In a binary mixture one of the rows containing composition is irrelevant thus it's deleted
            X = np.delete(X,0,0)
            self.X = X
        else:
            self.binary = False
        
        
        #Casting whatever may pop up to float64, (Some scan cases lead to string data values)
        if X.dtype == "<U32":
            X = np.char.replace(X, ' ', '')
            X = X.astype(float)
        if Y.dtype == "<U32":
            Y = np.char.replace(Y, ' ', '')
            Y = Y.astype(float)
        self.Y = Y
        self.X = X
        return 0
        
    #Creates a regression of the data
    def regressData(self):
        x = np.array(self.X[-1,:])
        y = np.array(self.Y)
        if self.propert.lower() == 'viscosity':
            basis_func = lambda x,a,b: self.Arrhenius(x,a,b)
            p0 = [0.2,2000]
            self.regressionType = 'a*exp(b/(8.314*T))'
        elif self.propert.lower() == 'density' or self.propert.lower() == 'thermal conductivity' or self.propert.lower() == 'surface tension':
            basis_func = lambda x,a,b: self.oneDimensionalLine(x,a,b)
            p0 = [0.2,2000]
            self.regressionType = 'a + bT'
        elif self.propert.lower() == 'electric conductance':
            basis_func = lambda x,a,b,c: self.parabolic(x,a,b,c)
            p0 = [0,0,0]
            self.regressionType = 'a + bT + cT^2'
        elif self.propert.lower() == 'heat capacity':
            basis_func = lambda x,a,b,c: self.revParabolic(x,a,b,c)
            p0 = [0,0,0]
            self.regressionType = 'a + bT + cT^(-2)'
        elif self.propert.lower() == 'vapor pressure':
            basis_func = lambda x,a,b: self.logfun(x,a,b)
            p0 = [0,0]
            self.regressionType = '10^(a + b/T)'
        params,covar = optimize.curve_fit(basis_func,x,y,p0)
        self.parameters = params
        return 0
    
    def buildModel(self):
        if self.modelType == 'Regression kriging':
            print('Statistical model: Regression kriging')
            X = self.X
            Y = self.Y

            if self.oneDimensional:
                X = X[-1,:]
                X = X.reshape(-1,1)
            else:
                X = X.transpose()
            k = regression_kriging(X,Y)
            k.train()
            self.fitModel = k
            
        else:
            print('Invalid model type request')
        return 0
    
    #Using matplotlib, this generates figures containing datapoint and regression
    def makePlot(self):
        #If only one composition exists
        if self.oneDimensional:
            
            
            #Create data regression line
            xData = np.array(self.X[-1,:])
            yData = np.array(self.Y)
            self.regressData()
            xx = np.linspace(min(xData),max(xData),num=100)
            if self.propert.lower() == 'viscosity':
                yy = self.Arrhenius(xx,self.parameters[0],self.parameters[1])
            elif self.propert.lower() == 'density' or self.propert.lower() == 'thermal conductivity' or self.propert.lower() == 'surface tension':
                yy = self.oneDimensionalLine(xx,self.parameters[0],self.parameters[1])
            elif self.propert.lower() == 'electric conductance':
                yy = self.parabolic(xx,self.parameters[0],self.parameters[1],self.parameters[2])
            elif self.propert.lower() == 'heat capacity':
                yy = self.revParabolic(xx,self.parameters[0],self.parameters[1],self.parameters[2])
            elif self.propert.lower() == 'vapor pressure':
                yy = self.logfun(xx,self.parameters[0],self.parameters[1])
            #Since graph is one dimensional we extract the composition to print on graph
            comp = np.zeros(self.X.shape[0]-1)
            
            for i in range(len(comp)):
                comp[i] = self.X[i,0]
            
            
            #Create blackbox fit line
            yFit = np.zeros(xx.shape)
            for i in range(len(xx)):
                yFit[i] = self.fitModel.predict([xx[i]])

            
            self.xx = xx
            self.yFit = yFit
            
            #First plot, data along with regression
            figRE = plt.figure(0)
            plt.xlabel('Temperature (°C)')
            plt.ylabel(self.propert.capitalize() + ' (' + self.unit + ')')
            plt.scatter(xData,yData)
            plt.plot(xx,yy)
            plt.legend(["Regression Fit","Data"])
            plt.title(' '.join(self.salts) + ' ' + np.array2string(comp))
            plt.savefig('Regression.png')
            self.figRe = figRE
            
            #Second plot, data along with blackbox model
            figBL = plt.figure(1)
            plt.xlabel('Temperature (°C)')
            plt.ylabel(self.propert.capitalize() + ' (' + self.unit + ')')
            plt.scatter(xData,yData)
            plt.plot(xx,yFit,'r')
            plt.legend(["Blackbox Fit","Data"])
            plt.title(' '.join(self.salts) + ' ' + np.array2string(comp))
            plt.savefig('Model.png')
            self.figBL = figBL
            
        elif self.binary:
            alpha = np.linspace(np.amin(self.X[0,:]),np.amax(self.X[0,:]),100)
            T = np.linspace(np.amin(self.X[1,:]),np.amax(self.X[1,:]),100)
            print('Plotting surface plot')
            X,Y = np.meshgrid(alpha,T)
            Z = np.zeros(X.shape)
            for i in range(alpha.shape[0]):
                for j in range(T.shape[0]):
                    Z[i,j] = self.fitModel.predict([X[i,j],Y[i,j]])
            fig = plt.figure()
            ax = fig.gca(projection = '3d')
            surf = ax.plot_surface(X,Y,Z, cmap=cm.coolwarm, linewidth=0)
            ax.set_zlim(np.amin(self.Y),np.amax(self.Y))
            ax.set_xlim(alpha[0],alpha[-1])
            ax.set_ylim(T[0],T[-1])
            ax.zaxis.set_major_locator(LinearLocator(10))
            ax.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
            fig.colorbar(surf, shrink = 0.5, aspect=5)
            if np.mean(alpha) > 1:
                perOrFrac = 'Percent'
            else:
                perOrFrac = 'Fraction'
            ax.set_xlabel('Mole ' + perOrFrac + ' of ' + self.salts[0])
            ax.set_ylabel('Temperature °C')
            ax.set_zlabel(self.propert.capitalize() + ' ' + self.unit)
            ax.set_title(' '.join(self.salts))
            plt.savefig('Model.png')
            plt.show()
        else:
            print("Can't make relevant figure for dimensions larger than 3")
            
        return 0
        
    
    #Function creates the output excel file
    def printExcelReport(self):
        wb = openpyxl.Workbook()
        wb.create_sheet(index = 0, title='Overview')
        wb.create_sheet(index = 1, title='Data Export')
        wb.create_sheet(index = 2, title='Regression Info')
        wb.create_sheet(index = 3, title = 'Model Fit')
        
        sheet = wb.get_sheet_by_name('Overview')
        sheet['A1'] = 'Property'
        
        sheet['B1'] = self.propert.capitalize()
        sheet['C1'] = self.unit
        sheet['A2'] = 'Salts'
        for i in range(len(self.salts)):
            sheet.cell(row = 2, column=i+2).value = self.salts[i]
        
        sheet['A3'] = 'Regerssion Data Available'
        if self.regressionType == 'None':
            sheet['B3'] = 'No'
        else:
            sheet['B3'] = 'Yes'
            
        sheet['A4'] = 'Model fit done'
        if self.fitModel == 'None':
            sheet['B4'] = 'No'
        else:
            sheet['B4'] = 'Yes'
        
        sheet['A5'] = 'Number of datapoints'
        sheet['B5'] = self.X.shape[1]
        
        sheet['A6'] = 'Temperature Range (C)'
        sheet['B6'] = np.amin(self.X[-1,:])
        sheet['C6'] = np.amax(self.X[-1,:])
        
        sheet = wb.get_sheet_by_name('Data Export')
        sheet['A1'] = 'X'
        
        
        if self.binary:
            sheet['A2'] = self.salts[0] + ' to ' + self.salts[1] + ' molar ratio'
            sheet['A3'] = 'Temperature °C'
            
            for i in range(self.X.shape[1]):
                sheet.cell(row = 2, column = i + 2).value = self.X[0,i]
                sheet.cell(row = 3, column = i + 2).value = self.X[1,i]
            i = 1
        
        else:
            for i in range(len(self.salts)):
                sheet.cell(row = i+2, column = 1).value = self.salts[i]
                sheet.cell(row = i+3, column = 1).value = 'Temperature °C'
            
        for i in range(self.X.shape[0]):
            for j in range(self.X.shape[1]):
                sheet.cell(row = i + 2, column =  j+2).value = self.X[i,j]
        
        yStart = i + 4
        sheet.cell(row = yStart, column = 1).value = 'Y'
        sheet.cell(row = yStart+1, column = 1).value = self.propert.capitalize()
        
        
        for i in range(self.Y.shape[0]):
            sheet.cell(row = yStart +1, column = i+2).value = self.Y[i]
        
        sheet = wb.get_sheet_by_name('Regression Info')
        
        if self.regressionType == 'None':
            sheet['A1'] = 'N/A'
        else:
            sheet['A1'] = 'Regression Function'
            sheet['B1'] = self.regressionType
            sheet['A2'] = 'Parameters'
            for i in range(len(self.parameters)):
                sheet.cell(row = 2, column = 2+i).value = self.parameters[i]
                
            imag = Image('Regression.png')
            sheet.add_image(imag,'A3')
            
        sheet = wb.get_sheet_by_name('Model Fit')
        if self.fitModel == 'None':
            sheet['A1'] = 'N/A'
        elif self.oneDimensional or self.binary:
            image = Image('Model.png')
            sheet.add_image(image,'A1')
        else:
            sheet['A1'] = 'Model details in Python API object'
        
        newLoc = self.loc + '\..\Output'
        wb.save('output.xlsx')
        
    
        
        return 0
    
    def initializeData(self):
        self.scanLibrary()
        self.organizeData()
        self.printExcelReport()
        
    def initializeFull(self):
        self.scanLibrary()
        self.organizeData()
        if self.oneDimensional:
            self.regressData()
        self.buildModel()
        self.makePlot()
        self.printExcelReport()
    
    
    ## Regression Base Functions
    #Arrhenius function, default for Viscosity Regression
    def Arrhenius(self,x,a,b):
        return a*np.exp(b/(8.314*x))
    
    #Simple linear regression for Density, Surface Tension and Thermal Conductivity
    def oneDimensionalLine(self,x,a,b):
        return a + b*x
    
    #Function for Eletrical Conductance fit
    def parabolic(self,x,a,b,c):
        return a + b*x + c*x**2
    
    #Function to fit Heat Capacity
    def revParabolic(self,x,a,b,c):
        return a + b*x + c*x**(-2)
    
    #Function to fit Vapor Pressure
    def logfun(self,x,a,b):
        d = a + b/x
        return 10**d
        
#Run code after this line
        
""" Example Run
binaryMixture = API(’density’,[’LiCl’,’KCl’])
binaryMixture.initializeFull()
print(binaryMixture.fitModel.predict([62.3,752.4]))
"""